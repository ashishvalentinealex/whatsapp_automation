from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Load Excel file
file_path = r"C:\Users\avati\Documents\E-CHURCH\WAContacts290324.xlsx"  # Replace with your Excel file path
workbook = load_workbook(file_path)
sheet = workbook.active

# Set up WebDriver
driver = webdriver.Chrome()
driver.get("https://web.whatsapp.com/")
driver.maximize_window()
input("Scan QR Code and press Enter...")

# Define the common message template
common_message = "Hello {name}, I pray that this year may be a year of Restoration, Renewal, and Endless Possibilities."

# Iterate through the rows in the Excel sheet
for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
    name = row[0]  # Assuming 'Name' is in the first column
    # Extract first name
    name_parts = name.split()
    first_name = (
        name_parts[1] if len(name_parts[0]) <= 2 else name_parts[0]
    ) if name_parts else name

    # Personalize the message
    personalized_message = common_message.format(name=first_name)

    try:
        search_box = driver.find_element(By.XPATH, "(//p[contains(@class, 'selectable-text')])[1]")
        search_box.send_keys(name)
        search_box.send_keys(Keys.ENTER)
        time.sleep(5)

        try:
            cancel_button = driver.find_element(By.XPATH, "//button[@aria-label='Cancel search']")
            cancel_button.click()
        except Exception as e:
            print(f"Error occurred while cancelling for '{name}': {e}")
            continue


        try:
            driver.find_element(By.XPATH, "(//p[contains(@class, 'selectable-text')])[2]").send_keys(
                personalized_message)
            time.sleep(5)

            driver.find_element(By.XPATH, ".//span[@data-icon='send']").click()

        except Exception as e:
            print(f"Contact '{name}' not found. Skipping... : {e}")
            continue


    except Exception as e:
        print(f"Error occurred while searching for '{name}': {e}")
        continue


# Close the driver
driver.close()
